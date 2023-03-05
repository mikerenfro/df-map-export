#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Make an Excel file of Dwarf Fortress map elevations from text files written from a DFHack Lua script.

@author: Mike Renfro

Given a series of text files generated from a DFHack Lua script at different
elevations:

- For each elevation:
    - read a line from the elevation file (contains 0 or 1 in each position)
    - insert each value from the line into a new workbook sheet
    - apply a color scale conditional format to the sheet
"""
import argparse
import glob
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import os
import re


def main(world, zoom, basedir, embark_elevation):
    """
    Convert a folder of elevation files into editable Excel format.

    Parameters
    ----------
    basedir : str
        Base directory containing folders with minimap screenshots
    world : str
        Folder in basedir containing minimap screenshots
    zoom : int
        Spreadsheet zoom level (in percent)
    embark_elevation : int
        Elevation of embark site (to set active worksheet)

    Returns
    -------
    None.

    """
    minimap_dict = {}
    map_size = (None, None)
    if not glob.glob(os.path.join(basedir, world, 'elevation-*.txt')):
        raise OSError('{0} contains no elevation-*.txt files'.format(os.path.join(basedir, world)))
    for filename in glob.glob(os.path.join(basedir, world, 'elevation-*.txt')):
        elevation_list = re.findall('elevation-([0-9-+]+)\.txt', filename)
        elevation = int(elevation_list[0])
        with open(filename, 'r') as f:
            lines = f.readlines()
            if map_size == (None, None):
                map_size = (len(lines), len(lines[0].strip()))
        minimap_dict[elevation] = lines

    wb = Workbook()
    sheet_first_row = 1
    sheet_last_row = sheet_first_row + map_size[0] - 1
    sheet_first_column = 1
    sheet_last_column = sheet_first_column + map_size[1] - 1
    sheet_first_column_letter = get_column_letter(sheet_first_column)
    sheet_last_column_letter = get_column_letter(sheet_last_column)
    print("Converting underground pixels in elevation: ", end='', flush=True)
    elevations = sorted(list(minimap_dict.keys()), key=int, reverse=True)
    # for elevation, minimap in minimap_dict.items():
    for elevation in elevations:
        minimap = minimap_dict[elevation]
        ws = wb.create_sheet(title="Elev {0}".format(elevation))
        print(" {0}".format(elevation), end='', flush=True)
        for row, line in enumerate(minimap, start=1):
            for col, pixel in enumerate(line.strip(), start=1):
                if (pixel=='1'):
                    # print('Should make a black pixel at (row, col)=({0}, {1})'.format(row, col))
                    _ = ws.cell(column=col, row=row, value=int(pixel))

        # Create conditional format rule for this sheet
        rule = ColorScaleRule(start_type='min', start_color='FFFFFF',
                              end_type='max', end_color='000000')
        ws.conditional_formatting.add('{0}{1}:{2}{3}'.format(
            sheet_first_column_letter, sheet_first_row,
            sheet_last_column_letter, sheet_last_row), rule)
        # Set width of columns
        for i in range(sheet_first_column, sheet_last_column+1):
            ws.column_dimensions[get_column_letter(i)].width = 2.875
        ws.sheet_view.zoomScale = zoom

    print(" done.")

    if not (embark_elevation is None):
        wb.active = wb['Elev {0}'.format(embark_elevation)]
    del wb['Sheet']
    print("Saving spreadsheet: ", end='', flush=True)
    wb.save("{0}.xlsx".format(world))
    print("done.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--basedir", type=str, default='elevations',
                        help="Base directory containing folders with minimap elevations (defaults to 'elevations')")
    parser.add_argument("--zoom", type=int, default=25,
                        help="Spreadsheet zoom level (in percent, defaults to 25)")
    parser.add_argument("--embark-elevation", type=int,
                        help="Elevation of embark site (if specified, will set active sheet in Excel workbook)")
    parser.add_argument("world", type=str,
                        help="Folder in basedir containing minimap screenshots")
    args = parser.parse_args()
    main(world=args.world, basedir=args.basedir, zoom=args.zoom, embark_elevation=args.embark_elevation)
