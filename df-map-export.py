#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Make an Excel file of Dwarf Fortress map elevations from text files written from a DFHack Lua script.

@author: Mike Renfro

Given a series of text files generated from a DFHack Lua script at different
elevations:

- For each elevation:
    - read a line from the elevation file (contains various characters in each position)
    - insert a blank cell into the worksheet for each position, with background color set by the character
"""
import argparse
import glob
import openpyxl
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os
import re
import string


def main(world, zoom, basedir, embark_elevation, enable_macros=False):
    """
    Convert a folder of elevation files into editable Excel format.

    Parameters
    ----------
    basedir : str
        Base directory containing folders with minimap screenshots
    world : str
        Folder in basedir containing minimap screenshots
    zoom : int
        Spreadsheet zoom level (in percent)
    embark_elevation : int
        Elevation of embark site (to set active worksheet)

    Returns
    -------
    None.

    """
    minimap_dict = {}
    map_size = (None, None)
    if not glob.glob(os.path.join(basedir, world, '*.txt')):
        raise OSError('{0} contains no elevation-*.txt files'.format(os.path.join(basedir, world)))
    for filename in glob.glob(os.path.join(basedir, world, '*.txt')):
        elevation_list = re.findall('-([0-9-+]+)\.txt$', filename)
        elevation = int(elevation_list[0])
        with open(filename, 'r') as f:
            lines = f.readlines()
            if map_size == (None, None):
                map_size = (len(lines), len(lines[0].rstrip('\n')))
        minimap_dict[elevation] = lines

    if enable_macros:
        wb = openpyxl.reader.excel.load_workbook('template.xlsm', keep_vba=True)
        filename = "{0}.xlsm".format(world)
    else:
        wb = Workbook()
        filename = "{0}.xlsx".format(world)

    sheet_first_column = 1
    sheet_last_column = sheet_first_column + map_size[1] - 1
    # print("map_size =", map_size)
    # print("sheet_last_column =", sheet_last_column)
    print("Converting underground pixels in elevation: ", end='', flush=True)
    elevations = sorted(list(minimap_dict.keys()), key=int, reverse=True)
    # Create styles for this sheet
    black_fill = PatternFill(start_color='010101',
                             end_color='010101',
                             fill_type='solid')
    light_brown_fill = PatternFill(start_color='C4A484',
                             end_color='C4A484',
                             fill_type='solid')
    dark_brown_fill = PatternFill(start_color='5C4033',
                                  end_color='5C4033',
                                  fill_type='solid')
    light_grey_fill = PatternFill(start_color='A6A6A6',
                            end_color='A6A6A6',
                            fill_type='solid')
    dark_grey_fill = PatternFill(start_color='5A5A5A',
                            end_color='5A5A5A',
                            fill_type='solid')
    light_green_fill = PatternFill(start_color='92D050',
                                   end_color='92D050',
                                   fill_type='solid')
    dark_green_fill = PatternFill(start_color='76933C',
                                  end_color='76933C',
                                  fill_type='solid')
    orange_fill = PatternFill(start_color='E26B0A',
                              end_color='E26B0A',
                              fill_type='solid')
    blue_fill = PatternFill(start_color='0000FF',
                            end_color='0000FF',
                            fill_type='solid')
    fill_dict = {
        '?': black_fill,
        'T': light_brown_fill,
        's': dark_brown_fill,
        'r': dark_grey_fill,
        'B': light_grey_fill,
        'g': light_green_fill,
        'p': dark_green_fill,
        'M': orange_fill,
        '~': blue_fill,
    }
    for elevation in elevations:
        minimap = minimap_dict[elevation]
        ws = wb.create_sheet(title="Elev {0}".format(elevation))
        print(" {0}".format(elevation), end='', flush=True)
        for row, line in enumerate(minimap, start=1):
            for col, pixel in enumerate(line.rstrip('\n'), start=1):
                if pixel == ' ':
                    continue
                else:
                    c = ws.cell(column=col, row=row)
                    if pixel in fill_dict:
                        c.fill = fill_dict[pixel]
        # Set width of columns
        for i in range(sheet_first_column, sheet_last_column+1):
            ws.column_dimensions[get_column_letter(i)].width = 2.875
        ws.sheet_view.zoomScale = zoom

    print(" done.")

    # del wb['Sheet']
    if not (embark_elevation is None):
        print('Setting Elev {0} as active'.format(embark_elevation))
        wb.active = wb['Elev {0}'.format(embark_elevation)]
    print("Saving spreadsheet: ", end='', flush=True)
    wb.save(filename)
    print("done.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--basedir", type=str, default='elevations',
                        help="Base directory containing folders with minimap elevations (defaults to 'elevations')")
    parser.add_argument("--zoom", type=int, default=25,
                        help="Spreadsheet zoom level (in percent, defaults to 25)")
    parser.add_argument("--embark-elevation", type=int,
                        help="Elevation of embark site (if specified, will set active sheet in Excel workbook)")
    parser.add_argument("--enable-macros", action='store_true',
                        help="Create Excel workbook with macros to propagate zoom level and cell location across sheets")
    parser.add_argument("world", type=str,
                        help="Folder in basedir containing minimap screenshots")
    args = parser.parse_args()
    main(world=args.world, basedir=args.basedir, zoom=args.zoom,
         embark_elevation=args.embark_elevation,
         enable_macros=args.enable_macros)
